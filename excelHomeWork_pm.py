# -*- coding: utf-8 -*-
# @Author:sunny
# @File :excelHomeWork_pm.py
# @Software :PyCharm
# @Time :2021/6/16 10:10

# 接口自动化的步骤
# 1.编写好测试用例，代码自动读取测试用例里的数据 read_data()
# 2.发送接口请求，得到响应结果 -- 实际结果 func()函数
# 3.执行结果 vs 预期结果
# 4.写入最终的测试结果到测试用例 --write_data()

import openpyxl
import requests


def func_post(url, data, headers=None):
    if headers is None:
        headers = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    res_post = requests.post(url=url, json=data, headers=headers).json()
    return res_post


def read_data(fileName, sheetName):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb[sheetName]  # 找到sheet
    row = sheet.max_row  # 获取最大的行数
    list_1 = []
    for item in range(2, row + 1):  # 取左不取右
        dict_1 = dict(
            id_reg=sheet.cell(row=item, column=1).value,  # 取id
            url_reg=sheet.cell(row=item, column=5).value,  # 取url
            data_reg=sheet.cell(row=item, column=6).value,  # 取data
            expected_reg=sheet.cell(row=item, column=7).value)  # 取data
        list_1.append(dict_1)
    return list_1


def write_data(fileName, sheetName, row, column, finalResult):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb[sheetName]
    sheet.cell(row=row, column=column).value = finalResult
    wb.save(fileName)


# res = read_data('test_case_api.xlsx', 'register')  # 调用读取函数读取注册接口测试用例


# print(res)
def execute_function(fileName, sheetName):
    res = read_data(fileName, sheetName)
    for case in res:
        case_id = case['id_reg']
        case_url = case.get('url_reg')
        case_data = case.get('data_reg')
        case_expect = case.get('expected_reg')
        # print(case_id, case_url, case_data, case_expect)
        case_data = eval(case_data)
        case_expect = eval(case_expect)
        real_result = func_post(url=case_url, data=case_data)
        # print(real_result)
        # print(case_data)
        # print(type(case_data))
        # dict_1 = {"aaa": "123123", "bbb": "123123123"}
        # print(dict_1)
        # print(type(dict_1))
        case_expect_msg = case_expect['msg']
        real_result_msg = real_result['msg']
        print('用例编号：{}'.format(case_id))
        print('预期结果为：{}'.format(case_expect_msg))
        print('实际结果为：{}'.format(real_result_msg))
        print(case_expect_msg, real_result_msg)
        if case_expect_msg == real_result_msg:
            print('这条用例通过')
            final_result = 'pass'
        else:
            print('这条用例不通过')
            final_result = 'fail'
        print('*' * 50)
        write_data(fileName, sheetName, case_id + 1, 8, final_result)


execute_function('test_case_api.xlsx', 'register')
execute_function('test_case_api.xlsx', 'login')
